import json
import random
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment

with open('/share/shmatikov/collin/adversarial_decoding/data/naturalness_decoding.json', 'r') as f:
    naturalness_decoding = json.load(f)

random.seed(42)
random.shuffle(naturalness_decoding)
res = [p['generation'] for p in naturalness_decoding]

form = ''

ll = []
for r in res:
    q = {
        'text': r,
        'readability': ' '
    }
    ll.append(q)


df = pd.DataFrame(ll)

file_path = 'readability_eval.xlsx'
df.to_excel(file_path)

# Load the workbook and access the sheet
wb = load_workbook(file_path)
ws = wb.active  # Get the active worksheet

# Apply wrap text to all cells
for row in ws.iter_rows():
    for cell in row:
        cell.alignment = Alignment(wrap_text=True)

# Save the modified workbook
wb.save(file_path)